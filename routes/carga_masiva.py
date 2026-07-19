from flask import Blueprint, render_template, request, flash, redirect, url_for, session, send_file
from models import get_db_connection
from werkzeug.security import generate_password_hash
from functools import wraps
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

carga_masiva_bp = Blueprint('carga_masiva_bp', __name__)


HEADER_FILL = PatternFill(start_color='2C7BE5', end_color='2C7BE5', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='2C3E50')
INFO_FILL = PatternFill(start_color='EAF3FF', end_color='EAF3FF', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('rol') != 'admin':
            flash('Acceso no autorizado', 'danger')
            return redirect(url_for('auth_bp.login'))
        return f(*args, **kwargs)
    return decorated_function


def _norm(value):
    if value is None or pd.isna(value):
        return ''
    return str(value).strip()


def _norm_upper(value):
    return _norm(value).upper()


def _safe_int(value):
    text = _norm(value)
    if not text:
        return None
    try:
        return int(float(text))
    except Exception:
        return None


def _username_slug(text):
    return ''.join(ch for ch in _norm(text).lower().replace(' ', '') if ch.isalnum() or ch == '_')


def _build_username(prefix, nombres, apellidos):
    return f"{prefix}_{_username_slug(nombres)}_{_username_slug(apellidos)}"


def _read_uploaded_dataframe(archivo):
    if archivo.filename.lower().endswith('.csv'):
        return pd.read_csv(archivo)
    return pd.read_excel(archivo)


def _resolver_grupo(cursor, row):
    id_grupo = _safe_int(row.get('ID_Grupo'))
    if id_grupo is not None:
        cursor.execute('SELECT ID_Grupo, Nombre_Grupo FROM grupos WHERE ID_Grupo = %s', (id_grupo,))
        return cursor.fetchone()

    nombre = _norm(row.get('Nombre_Grupo') or row.get('Grupo'))
    if nombre:
        cursor.execute('SELECT ID_Grupo, Nombre_Grupo FROM grupos WHERE UPPER(Nombre_Grupo) = %s', (_norm_upper(nombre),))
        return cursor.fetchone()
    return None


def _resolver_materia(cursor, row):
    id_materia = _safe_int(row.get('ID_Materia'))
    if id_materia is not None:
        cursor.execute('SELECT ID_Materia, Nombre_Materia FROM materias WHERE ID_Materia = %s', (id_materia,))
        return cursor.fetchone()

    nombre = _norm(row.get('Nombre_Materia') or row.get('Materia'))
    if nombre:
        cursor.execute('SELECT ID_Materia, Nombre_Materia FROM materias WHERE UPPER(Nombre_Materia) = %s', (_norm_upper(nombre),))
        return cursor.fetchone()
    return None


def _resolver_docente(cursor, row):
    id_docente = _safe_int(row.get('ID_Docente'))
    if id_docente is not None:
        cursor.execute(
            'SELECT d.ID_Docente, d.Nombres, d.Apellidos, u.Usuario FROM docentes d LEFT JOIN usuarios u ON d.ID_Usuario = u.ID_Usuario WHERE d.ID_Docente = %s',
            (id_docente,),
        )
        return cursor.fetchone()

    usuario = _norm(row.get('Usuario_Docente') or row.get('Usuario'))
    if usuario:
        cursor.execute(
            '''SELECT d.ID_Docente, d.Nombres, d.Apellidos, u.Usuario
               FROM docentes d
               JOIN usuarios u ON d.ID_Usuario = u.ID_Usuario
               WHERE UPPER(u.Usuario) = %s''',
            (_norm_upper(usuario),),
        )
        return cursor.fetchone()

    nombre_completo = _norm(row.get('Docente') or row.get('Docente_Nombre'))
    if nombre_completo:
        cursor.execute(
            '''SELECT d.ID_Docente, d.Nombres, d.Apellidos, u.Usuario
               FROM docentes d
               LEFT JOIN usuarios u ON d.ID_Usuario = u.ID_Usuario
               WHERE UPPER(TRIM(d.Nombres || ' ' || d.Apellidos)) = %s''',
            (_norm_upper(nombre_completo),),
        )
        return cursor.fetchone()

    nombres = _norm(row.get('Nombres'))
    apellidos = _norm(row.get('Apellidos'))
    if nombres and apellidos:
        cursor.execute(
            '''SELECT d.ID_Docente, d.Nombres, d.Apellidos, u.Usuario
               FROM docentes d
               LEFT JOIN usuarios u ON d.ID_Usuario = u.ID_Usuario
               WHERE UPPER(d.Nombres) = %s AND UPPER(d.Apellidos) = %s''',
            (_norm_upper(nombres), _norm_upper(apellidos)),
        )
        return cursor.fetchone()
    return None


def _format_main_sheet(ws, title, headers):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws.freeze_panes = 'A3'
    for idx, header in enumerate(headers, start=1):
        width = max(16, min(32, len(str(header)) + 4))
        ws.column_dimensions[get_column_letter(idx)].width = width


def _add_instructions_sheet(wb, title, lines):
    ws = wb.create_sheet('INSTRUCCIONES')
    ws.append(['Instrucción'])
    for line in lines:
        ws.append([line])
    _format_main_sheet(ws, title, ['Instrucción'])
    ws.column_dimensions['A'].width = 120
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')


def _add_reference_sheet(wb, title, headers, rows, hidden=False):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    _format_main_sheet(ws, f'📎 REFERENCIAS - {title}', headers)
    if hidden:
        ws.sheet_state = 'hidden'
    return ws


def _add_dropdown_validation(ws, target_range, source_sheet, source_col_letter, start_row, end_row):
    if end_row < start_row:
        return
    formula = f"='{source_sheet}'!${source_col_letter}${start_row}:${source_col_letter}${end_row}"
    dv = DataValidation(type='list', formula1=formula, allow_blank=True)
    dv.prompt = 'Seleccione un valor de la lista'
    dv.error = 'Seleccione un valor válido de la lista desplegable'
    ws.add_data_validation(dv)
    dv.add(target_range)


@carga_masiva_bp.route('/carga_masiva')
@admin_required
def carga_masiva():
    return render_template('carga_masiva.html')


@carga_masiva_bp.route('/carga_masiva/estudiantes', methods=['POST'])
@admin_required
def carga_masiva_estudiantes():
    if 'archivo' not in request.files or request.files['archivo'].filename == '':
        flash('No se seleccionó ningún archivo', 'danger')
        return redirect(url_for('carga_masiva_bp.carga_masiva'))

    archivo = request.files['archivo']
    try:
        df = _read_uploaded_dataframe(archivo)
        columnas_requeridas = {'Nombres', 'Apellidos'}
        faltantes = [c for c in columnas_requeridas if c not in df.columns]
        if faltantes:
            flash(f'Faltan columnas requeridas: {", ".join(faltantes)}', 'danger')
            return redirect(url_for('carga_masiva_bp.carga_masiva'))

        conn = get_db_connection(); cursor = conn.cursor()
        exitosos = 0; errores = []; omitidos = 0

        for idx, row in df.iterrows():
            try:
                nombres = _norm(row.get('Nombres'))
                apellidos = _norm(row.get('Apellidos'))
                grupo = _resolver_grupo(cursor, row)
                if not nombres or not apellidos:
                    errores.append(f'Fila {idx+2}: Nombres y apellidos son obligatorios')
                    continue
                if not grupo:
                    errores.append(f'Fila {idx+2}: Grupo no válido')
                    continue

                cursor.execute(
                    'SELECT ID_Estudiante FROM estudiantes WHERE UPPER(Nombres)=%s AND UPPER(Apellidos)=%s AND ID_Grupo=%s',
                    (_norm_upper(nombres), _norm_upper(apellidos), grupo['ID_Grupo']),
                )
                if cursor.fetchone():
                    omitidos += 1
                    continue

                usuario = _norm(row.get('Usuario')) or _build_username('est', nombres, apellidos)
                cursor.execute(
                    'INSERT INTO estudiantes (Nombres, Apellidos, ID_Grupo, ID_Usuario) VALUES (%s, %s, %s, NULL)',
                    (nombres, apellidos, grupo['ID_Grupo']),
                )
                exitosos += 1
            except Exception as e:
                errores.append(f'Fila {idx+2}: {str(e)}')

        conn.commit(); conn.close()
        flash(f'✅ {exitosos} estudiantes cargados | ⏭️ {omitidos} omitidos | ❌ {len(errores)} errores', 'success' if not errores else 'warning')
        for error in errores[:8]:
            flash(error, 'danger')
    except Exception as e:
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')
    return redirect(url_for('carga_masiva_bp.carga_masiva'))


@carga_masiva_bp.route('/carga_masiva/docentes', methods=['POST'])
@admin_required
def carga_masiva_docentes():
    if 'archivo' not in request.files or request.files['archivo'].filename == '':
        flash('No se seleccionó ningún archivo', 'danger')
        return redirect(url_for('carga_masiva_bp.carga_masiva'))

    archivo = request.files['archivo']
    try:
        df = _read_uploaded_dataframe(archivo)
        columnas_requeridas = {'Nombres', 'Apellidos'}
        faltantes = [c for c in columnas_requeridas if c not in df.columns]
        if faltantes:
            flash(f'Faltan columnas requeridas: {", ".join(faltantes)}', 'danger')
            return redirect(url_for('carga_masiva_bp.carga_masiva'))

        conn = get_db_connection(); cursor = conn.cursor()
        exitosos = 0; errores = []; omitidos = 0
        for idx, row in df.iterrows():
            try:
                nombres = _norm(row.get('Nombres')); apellidos = _norm(row.get('Apellidos'))
                if not nombres or not apellidos:
                    errores.append(f'Fila {idx+2}: Nombres y apellidos son obligatorios')
                    continue

                cursor.execute('SELECT ID_Docente FROM docentes WHERE UPPER(Nombres)=%s AND UPPER(Apellidos)=%s', (_norm_upper(nombres), _norm_upper(apellidos)))
                if cursor.fetchone():
                    omitidos += 1
                    continue

                usuario = _norm(row.get('Usuario')) or _build_username('doc', nombres, apellidos)
                cursor.execute('SELECT ID_Usuario FROM usuarios WHERE Usuario = %s', (usuario,))
                user = cursor.fetchone()
                if user:
                    id_usuario = user['ID_Usuario']
                else:
                    cursor.execute('INSERT INTO usuarios (Usuario, Contrasena, Rol) VALUES (%s, %s, %s)', (usuario, generate_password_hash('123456'), 'docente'))
                    id_usuario = cursor.lastrowid

                cursor.execute('INSERT INTO docentes (Nombres, Apellidos, ID_Usuario) VALUES (%s, %s, %s)', (nombres, apellidos, id_usuario))
                exitosos += 1
            except Exception as e:
                errores.append(f'Fila {idx+2}: {str(e)}')
        conn.commit(); conn.close()
        flash(f'✅ {exitosos} docentes cargados | ⏭️ {omitidos} omitidos | ❌ {len(errores)} errores', 'success' if not errores else 'warning')
        for error in errores[:8]:
            flash(error, 'danger')
    except Exception as e:
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')
    return redirect(url_for('carga_masiva_bp.carga_masiva'))


@carga_masiva_bp.route('/carga_masiva/asignaturas', methods=['POST'])
@admin_required
def carga_masiva_asignaturas():
    if 'archivo' not in request.files or request.files['archivo'].filename == '':
        flash('No se seleccionó ningún archivo', 'danger')
        return redirect(url_for('carga_masiva_bp.carga_masiva'))

    archivo = request.files['archivo']
    try:
        df = _read_uploaded_dataframe(archivo)
        if 'Nombre_Materia' not in df.columns:
            flash('Falta la columna requerida: Nombre_Materia', 'danger')
            return redirect(url_for('carga_masiva_bp.carga_masiva'))

        conn = get_db_connection(); cursor = conn.cursor()
        exitosos = 0; errores = []; omitidos = 0
        for idx, row in df.iterrows():
            try:
                nombre_materia = _norm(row.get('Nombre_Materia'))
                if not nombre_materia:
                    errores.append(f'Fila {idx+2}: Nombre_Materia es obligatorio')
                    continue
                es_dim = _safe_int(row.get('Es_DIM'))
                es_dim = 1 if es_dim == 1 or nombre_materia.upper().startswith('DIM') else 0
                id_area = None
                nombre_area = _norm(row.get('Nombre_Area'))
                if nombre_area:
                    cursor.execute('SELECT ID_Area FROM areas WHERE UPPER(Nombre_Area)=%s', (_norm_upper(nombre_area),))
                    area = cursor.fetchone()
                    if area:
                        id_area = area['ID_Area']
                cursor.execute('SELECT ID_Materia FROM materias WHERE UPPER(Nombre_Materia)=%s', (_norm_upper(nombre_materia),))
                if cursor.fetchone():
                    omitidos += 1
                    continue
                cursor.execute('INSERT INTO materias (Nombre_Materia, ID_Area, Es_DIM) VALUES (%s, %s, %s)', (nombre_materia, id_area, es_dim))
                exitosos += 1
            except Exception as e:
                errores.append(f'Fila {idx+2}: {str(e)}')
        conn.commit(); conn.close()
        flash(f'✅ {exitosos} asignaturas cargadas | ⏭️ {omitidos} omitidas | ❌ {len(errores)} errores', 'success' if not errores else 'warning')
        for error in errores[:8]:
            flash(error, 'danger')
    except Exception as e:
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')
    return redirect(url_for('carga_masiva_bp.carga_masiva'))


@carga_masiva_bp.route('/carga_masiva/plan_estudios', methods=['POST'])
@admin_required
def carga_masiva_plan_estudios():
    if 'archivo' not in request.files or request.files['archivo'].filename == '':
        flash('No se seleccionó ningún archivo', 'danger')
        return redirect(url_for('carga_masiva_bp.carga_masiva'))

    archivo = request.files['archivo']
    try:
        df = _read_uploaded_dataframe(archivo)
        conn = get_db_connection(); cursor = conn.cursor()
        exitosos = 0; errores = []; omitidos = 0
        for idx, row in df.iterrows():
            try:
                grupo = _resolver_grupo(cursor, row)
                materia = _resolver_materia(cursor, row)
                if not grupo or not materia:
                    errores.append(f'Fila {idx+2}: Grupo o materia no válidos')
                    continue
                cursor.execute('SELECT ID_Plan FROM plan_estudios WHERE ID_Grupo=%s AND ID_Materia=%s', (grupo['ID_Grupo'], materia['ID_Materia']))
                if cursor.fetchone():
                    omitidos += 1
                    continue
                cursor.execute('INSERT INTO plan_estudios (ID_Grupo, ID_Materia) VALUES (%s, %s)', (grupo['ID_Grupo'], materia['ID_Materia']))
                exitosos += 1
            except Exception as e:
                errores.append(f'Fila {idx+2}: {str(e)}')
        conn.commit(); conn.close()
        flash(f'✅ {exitosos} materias agregadas al plan de estudios | ⏭️ {omitidos} omitidas | ❌ {len(errores)} errores', 'success' if not errores else 'warning')
        for error in errores[:8]:
            flash(error, 'danger')
    except Exception as e:
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')
    return redirect(url_for('carga_masiva_bp.carga_masiva'))


@carga_masiva_bp.route('/carga_masiva/carga_academica', methods=['POST'])
@admin_required
def carga_masiva_carga_academica():
    if 'archivo' not in request.files or request.files['archivo'].filename == '':
        flash('No se seleccionó ningún archivo', 'danger')
        return redirect(url_for('carga_masiva_bp.carga_masiva'))

    archivo = request.files['archivo']
    try:
        df = _read_uploaded_dataframe(archivo)
        conn = get_db_connection(); cursor = conn.cursor()
        creados = 0; actualizados = 0; omitidos = 0; errores = []
        for idx, row in df.iterrows():
            try:
                docente = _resolver_docente(cursor, row)
                grupo = _resolver_grupo(cursor, row)
                materia = _resolver_materia(cursor, row)
                if not docente or not grupo or not materia:
                    errores.append(f'Fila {idx+2}: Docente, grupo o materia no válidos')
                    continue

                cursor.execute('SELECT ID_Plan FROM plan_estudios WHERE ID_Grupo=%s AND ID_Materia=%s', (grupo['ID_Grupo'], materia['ID_Materia']))
                if not cursor.fetchone():
                    errores.append(f'Fila {idx+2}: La materia no está incluida en el plan de estudios del grupo')
                    continue

                cursor.execute('SELECT ID_Carga, ID_Docente FROM cargas_academicas WHERE ID_Grupo=%s AND ID_Materia=%s', (grupo['ID_Grupo'], materia['ID_Materia']))
                existente = cursor.fetchone()
                if existente:
                    if existente['ID_Docente'] == docente['ID_Docente']:
                        omitidos += 1
                    else:
                        cursor.execute('UPDATE cargas_academicas SET ID_Docente=%s WHERE ID_Carga=%s', (docente['ID_Docente'], existente['ID_Carga']))
                        actualizados += 1
                    continue

                cursor.execute('INSERT INTO cargas_academicas (ID_Docente, ID_Materia, ID_Grupo) VALUES (%s, %s, %s)', (docente['ID_Docente'], materia['ID_Materia'], grupo['ID_Grupo']))
                creados += 1
            except Exception as e:
                errores.append(f'Fila {idx+2}: {str(e)}')
        conn.commit(); conn.close()
        flash(f'✅ {creados} cargas creadas | 🔄 {actualizados} reasignadas | ⏭️ {omitidos} sin cambios | ❌ {len(errores)} errores', 'success' if not errores else 'warning')
        for error in errores[:8]:
            flash(error, 'danger')
    except Exception as e:
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')
    return redirect(url_for('carga_masiva_bp.carga_masiva'))


def _catalog_data():
    conn = get_db_connection(); cursor = conn.cursor()
    cursor.execute("SELECT ID_Grupo, Nombre_Grupo FROM grupos ORDER BY Nombre_Grupo")
    grupos = [(r['ID_Grupo'], r['Nombre_Grupo']) for r in cursor.fetchall()]
    cursor.execute("SELECT ID_Materia, Nombre_Materia FROM materias ORDER BY Nombre_Materia")
    materias = [(r['ID_Materia'], r['Nombre_Materia']) for r in cursor.fetchall()]
    cursor.execute("SELECT d.ID_Docente, TRIM(d.Nombres || ' ' || d.Apellidos) AS Docente, u.Usuario FROM docentes d LEFT JOIN usuarios u ON d.ID_Usuario = u.ID_Usuario ORDER BY d.Apellidos, d.Nombres")
    docentes = [(r['ID_Docente'], r['Docente'], r['Usuario'] or '') for r in cursor.fetchall()]
    cursor.execute("SELECT ID_Area, Nombre_Area FROM areas ORDER BY Nombre_Area")
    areas = [(r['ID_Area'], r['Nombre_Area']) for r in cursor.fetchall()]
    conn.close()
    return grupos, materias, docentes, areas


def _build_template(main_sheet, title, headers, sample_rows, instructions, references_builder=None, validations_builder=None):
    wb = Workbook()
    ws = wb.active
    ws.title = main_sheet
    ws.append(headers)
    for row in sample_rows:
        ws.append(row)
    _format_main_sheet(ws, title, headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=3, column=col).fill = INFO_FILL
    if references_builder:
        refs = references_builder(wb)
    else:
        refs = {}
    if validations_builder:
        validations_builder(wb, ws, refs)
    _add_instructions_sheet(wb, '📖 INSTRUCCIONES', instructions)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@carga_masiva_bp.route('/carga_masiva/plantilla_estudiantes')
@admin_required
def plantilla_estudiantes():
    grupos, _, _, _ = _catalog_data()
    def refs(wb):
        ws = _add_reference_sheet(wb, 'REFERENCIAS', ['ID_Grupo', 'Nombre_Grupo'], grupos)
        return {'grupos_end': ws.max_row}
    def vals(wb, ws, refs_dict):
        _add_dropdown_validation(ws, 'C3:C300', 'REFERENCIAS', 'B', 3, refs_dict['grupos_end'])
    output = _build_template(
        'Estudiantes',
        '📚 PLANTILLA DE ESTUDIANTES',
        ['Nombres', 'Apellidos', 'Grupo', 'Usuario'],
        [
            ['Juan', 'Pérez', grupos[0][1] if grupos else '', ''],
            ['María', 'González', grupos[1][1] if len(grupos) > 1 else '', ''],
        ],
        [
            'Use la columna Grupo con el nombre exacto del grupo o cargue luego por ID si prefiere CSV.',
            'La columna Usuario es opcional. Si la deja vacía, el sistema lo genera automáticamente.',
            'Contraseña inicial por defecto: 123456.',
            'No cambie los encabezados.',
        ],
        refs,
        vals,
    )
    return send_file(output, as_attachment=True, download_name='plantilla_estudiantes.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@carga_masiva_bp.route('/carga_masiva/plantilla_docentes')
@admin_required
def plantilla_docentes():
    output = _build_template(
        'Docentes',
        '👨‍🏫 PLANTILLA DE DOCENTES',
        ['Nombres', 'Apellidos', 'Usuario'],
        [['Ana', 'Martínez', ''], ['Luis', 'Sánchez', '']],
        [
            'La columna Usuario es opcional.',
            'Si no informa el usuario, el sistema crea uno tipo doc_nombre_apellido.',
            'Contraseña inicial por defecto: 123456.',
            'No cambie los encabezados.',
        ],
    )
    return send_file(output, as_attachment=True, download_name='plantilla_docentes.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@carga_masiva_bp.route('/carga_masiva/plantilla_asignaturas')
@admin_required
def plantilla_asignaturas():
    _, _, _, areas = _catalog_data()
    def refs(wb):
        ws = _add_reference_sheet(wb, 'REFERENCIAS', ['ID_Area', 'Nombre_Area'], areas)
        return {'areas_end': ws.max_row}
    def vals(wb, ws, refs_dict):
        _add_dropdown_validation(ws, 'B3:B300', 'REFERENCIAS', 'B', 3, refs_dict['areas_end'])
    output = _build_template(
        'Asignaturas',
        '📚 PLANTILLA DE ASIGNATURAS',
        ['Nombre_Materia', 'Nombre_Area', 'Es_DIM'],
        [['MATEMÁTICAS', areas[0][1] if areas else '', 0], ['DIM. COGNITIVA', 'COGNITIVA', 1]],
        [
            'Es_DIM es opcional: use 1 para preescolar y 0 para materia normal.',
            'Si el nombre empieza por DIM., el sistema también la marcará como DIM.',
            'Nombre_Area puede escogerse desde la lista.',
        ],
        refs,
        vals,
    )
    return send_file(output, as_attachment=True, download_name='plantilla_asignaturas.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@carga_masiva_bp.route('/carga_masiva/plantilla_plan_estudios')
@admin_required
def plantilla_plan_estudios():
    grupos, materias, _, _ = _catalog_data()
    def refs(wb):
        ws1 = _add_reference_sheet(wb, 'REF_GRUPOS', ['ID_Grupo', 'Nombre_Grupo'], grupos)
        ws2 = _add_reference_sheet(wb, 'REF_MATERIAS', ['ID_Materia', 'Nombre_Materia'], materias)
        return {'grupos_end': ws1.max_row, 'materias_end': ws2.max_row}
    def vals(wb, ws, refs_dict):
        _add_dropdown_validation(ws, 'A3:A500', 'REF_GRUPOS', 'B', 3, refs_dict['grupos_end'])
        _add_dropdown_validation(ws, 'B3:B500', 'REF_MATERIAS', 'B', 3, refs_dict['materias_end'])
    output = _build_template(
        'PlanEstudios',
        '📘 PLANTILLA DE PLAN DE ESTUDIOS',
        ['Grupo', 'Materia'],
        [[grupos[0][1] if grupos else '', materias[0][1] if materias else '']],
        [
            'Esta plantilla SOLO crea la relación materia-grupo del plan de estudios.',
            'Ya no asigna docentes por defecto.',
            'Use los nombres del grupo y la materia desde las listas desplegables.',
        ],
        refs,
        vals,
    )
    return send_file(output, as_attachment=True, download_name='plantilla_plan_estudios.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@carga_masiva_bp.route('/carga_masiva/plantilla_carga_academica')
@admin_required
def plantilla_carga_academica():
    grupos, materias, docentes, _ = _catalog_data()
    def refs(wb):
        ws1 = _add_reference_sheet(wb, 'REF_DOCENTES', ['ID_Docente', 'Docente', 'Usuario'], docentes)
        ws2 = _add_reference_sheet(wb, 'REF_GRUPOS', ['ID_Grupo', 'Nombre_Grupo'], grupos)
        ws3 = _add_reference_sheet(wb, 'REF_MATERIAS', ['ID_Materia', 'Nombre_Materia'], materias)
        return {'docentes_end': ws1.max_row, 'grupos_end': ws2.max_row, 'materias_end': ws3.max_row}
    def vals(wb, ws, refs_dict):
        _add_dropdown_validation(ws, 'A3:A500', 'REF_DOCENTES', 'B', 3, refs_dict['docentes_end'])
        _add_dropdown_validation(ws, 'B3:B500', 'REF_GRUPOS', 'B', 3, refs_dict['grupos_end'])
        _add_dropdown_validation(ws, 'C3:C500', 'REF_MATERIAS', 'B', 3, refs_dict['materias_end'])
    output = _build_template(
        'CargaAcademica',
        '🧑‍🏫 PLANTILLA DE CARGA ACADÉMICA',
        ['Docente', 'Grupo', 'Materia'],
        [[docentes[0][1] if docentes else '', grupos[0][1] if grupos else '', materias[0][1] if materias else '']],
        [
            'Use esta plantilla para asignar o reasignar el docente de una materia en un grupo.',
            'Si la materia ya tenía docente en ese grupo, el sistema actualizará la asignación.',
            'La materia debe existir primero en el plan de estudios del grupo.',
        ],
        refs,
        vals,
    )
    return send_file(output, as_attachment=True, download_name='plantilla_carga_academica.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
