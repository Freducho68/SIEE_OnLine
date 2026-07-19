from flask import Blueprint, render_template, request, flash, redirect, url_for, session, send_file
from models import get_db_connection, get_periodo_activo
from functools import wraps
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

planillas_bp = Blueprint('planillas_bp', __name__)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Debe iniciar sesión para acceder a esta página', 'warning')
            return redirect(url_for('auth_bp.login'))
        return f(*args, **kwargs)
    return decorated_function

def docente_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('rol') != 'docente':
            flash('Acceso no autorizado', 'danger')
            return redirect(url_for('auth_bp.panel_docente'))
        return f(*args, **kwargs)
    return decorated_function

def get_docente_id_from_user(cursor, user_id):
    cursor.execute('SELECT ID_Docente FROM docentes WHERE ID_Usuario = %s', (user_id,))
    docente = cursor.fetchone()
    return docente['ID_Docente'] if docente else None

def es_materia_dim(nombre_materia):
    if nombre_materia:
        nombre_upper = nombre_materia.upper()
        return nombre_upper.startswith('DIM') or nombre_upper.startswith('DIM.')
    return False

def validar_nota(nota, es_dim):
    if nota is None or nota == '':
        return True

    try:
        nota_float = float(nota)
        if es_dim:
            return nota_float in [3.0, 4.0, 5.0]
        return 0 <= nota_float <= 5
    except (ValueError, TypeError):
        return False

def calcular_promedio(notas):
    notas_validas = [n for n in notas if n is not None and n != '']
    if not notas_validas:
        return None
    try:
        promedio = sum(notas_validas) / len(notas_validas)
        return round(promedio, 1)
    except Exception:
        return None

def extraer_indicadores(texto):
    indicadores = ['' for _ in range(10)]
    if texto:
        lista = [item.strip() for item in str(texto).split('\n') if item and item.strip()]
        for i, valor in enumerate(lista[:10], start=1):
            indicadores[i - 1] = valor
    return indicadores

def unir_indicadores(valores):
    return '\n'.join([v.strip() for v in valores if v and str(v).strip()])

def obtener_indicadores_planilla(cursor, id_carga, periodo):
    cursor.execute(
        "SELECT Indicadores FROM indicadores_planilla WHERE ID_Carga = %s AND Periodo = %s",
        (id_carga, periodo),
    )
    fila = cursor.fetchone()
    return extraer_indicadores(fila['Indicadores']) if fila and fila['Indicadores'] else ['' for _ in range(10)]

def guardar_indicadores_planilla(cursor, id_carga, periodo, indicadores_lista):
    indicadores_texto = unir_indicadores(indicadores_lista)
    cursor.execute(
        """
        INSERT INTO indicadores_planilla (ID_Carga, Periodo, Indicadores)
        VALUES (%s, %s, %s)
        ON CONFLICT(ID_Carga, Periodo) DO UPDATE SET Indicadores = excluded.Indicadores
        """,
        (id_carga, periodo, indicadores_texto),
    )

def obtener_observador_periodo(cursor, id_estudiante, periodo):
    cursor.execute(
        """
        SELECT Observacion, Fallas
        FROM observador_estudiante
        WHERE ID_Estudiante = %s AND Periodo = %s
        """,
        (id_estudiante, periodo),
    )
    fila = cursor.fetchone()
    if fila:
        return {
            'observacion': fila['Observacion'] or '',
            'fallas': fila['Fallas'] or 0
        }
    return {'observacion': '', 'fallas': 0}

def guardar_observador_periodo(cursor, id_estudiante, periodo, observacion, fallas):
    cursor.execute(
        """
        INSERT INTO observador_estudiante (ID_Estudiante, Periodo, Observacion, Fallas)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT(ID_Estudiante, Periodo)
        DO UPDATE SET
            Observacion = excluded.Observacion,
            Fallas = excluded.Fallas,
            Fecha_Registro = CURRENT_TIMESTAMP
        """,
        (id_estudiante, periodo, observacion, fallas),
    )

def aplicar_formato_excel(writer, sheet_name):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color="2c7be5", end_color="2c7be5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 40)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    worksheet.freeze_panes = 'A2'


@planillas_bp.route('/planilla/<int:id_carga>', methods=['GET', 'POST'])
@login_required
@docente_required
def planilla(id_carga):
    conn = get_db_connection()
    cursor = conn.cursor()

    id_docente = get_docente_id_from_user(cursor, session['user_id'])
    if not id_docente:
        conn.close()
        flash('No se encontró información del docente', 'danger')
        return redirect(url_for('auth_bp.panel_docente'))

    cursor.execute(
        """
        SELECT c.ID_Carga, c.ID_Grupo, c.ID_Materia,
               m.Nombre_Materia, g.Nombre_Grupo, g.ID_Nivel
        FROM cargas_academicas c
        JOIN materias m ON c.ID_Materia = m.ID_Materia
        JOIN grupos g ON c.ID_Grupo = g.ID_Grupo
        WHERE c.ID_Carga = %s AND c.ID_Docente = %s
        """,
        (id_carga, id_docente)
    )
    info = cursor.fetchone()

    if not info:
        conn.close()
        flash('No tiene acceso a esta planilla', 'danger')
        return redirect(url_for('auth_bp.panel_docente'))

    is_dim = es_materia_dim(info['Nombre_Materia'])
    periodo_activo = get_periodo_activo()

    cursor.execute(
        """
        SELECT ID_Estudiante, Nombres, Apellidos
        FROM estudiantes
        WHERE ID_Grupo = %s
        ORDER BY Apellidos, Nombres
        """,
        (info['ID_Grupo'],)
    )

    estudiantes_rows = cursor.fetchall()
    indicadores_grupo = obtener_indicadores_planilla(cursor, id_carga, periodo_activo)
    estudiantes = []

    for estudiante_row in estudiantes_rows:
        estudiante = dict(estudiante_row)

        obs = obtener_observador_periodo(cursor, estudiante['ID_Estudiante'], periodo_activo)
        estudiante['observacion_periodo'] = obs['observacion']
        estudiante['fallas_periodo'] = obs['fallas']

        for i in range(1, 11):
            estudiante[f'eval_{i}'] = None
            estudiante[f'tarea_{i}'] = None

        cursor.execute(
            """
            SELECT Tipo, Numero, Nota
            FROM notas
            WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s
            """,
            (estudiante['ID_Estudiante'], id_carga, periodo_activo)
        )

        notas_existentes = cursor.fetchall()
        for nota in notas_existentes:
            tipo = nota['Tipo']
            numero = nota['Numero']
            valor = float(nota['Nota']) if nota['Nota'] is not None else None
            estudiante[f'{tipo}_{numero}'] = valor

        notas_eval = [estudiante[f'eval_{i}'] for i in range(1, 11)]
        notas_tarea = [estudiante[f'tarea_{i}'] for i in range(1, 11)]

        promedio_eval = calcular_promedio(notas_eval)
        promedio_tarea = calcular_promedio(notas_tarea)

        if is_dim:
            if promedio_eval is not None and promedio_tarea is not None:
                nota_final = (promedio_eval * 0.95) + (promedio_tarea * 0.05)
            elif promedio_eval is not None:
                nota_final = promedio_eval
            elif promedio_tarea is not None:
                nota_final = promedio_tarea
            else:
                nota_final = None

            if nota_final is not None:
                nota_final_redondeada = round(nota_final)
                if nota_final_redondeada < 3:
                    nota_final = 3.0
                elif nota_final_redondeada > 5:
                    nota_final = 5.0
                else:
                    nota_final = float(nota_final_redondeada)
        else:
            if promedio_eval is not None and promedio_tarea is not None:
                nota_final = round((promedio_eval * 0.95) + (promedio_tarea * 0.05), 1)
            elif promedio_eval is not None:
                nota_final = round(promedio_eval, 1)
            elif promedio_tarea is not None:
                nota_final = round(promedio_tarea, 1)
            else:
                nota_final = None

        estudiante['C1'] = nota_final
        estudiantes.append(estudiante)

    if request.method == 'POST':
        try:
            for estudiante in estudiantes:
                id_est = estudiante['ID_Estudiante']

                for i in range(1, 11):
                    campo_eval = f'eval_{id_est}_{i}'
                    valor_eval = request.form.get(campo_eval)

                    if valor_eval is not None and valor_eval != '':
                        try:
                            nota_valor = float(valor_eval)
                            if validar_nota(nota_valor, is_dim):
                                cursor.execute(
                                    """
                                    SELECT ID_Nota FROM notas
                                    WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'eval' AND Numero = %s
                                    """,
                                    (id_est, id_carga, periodo_activo, i)
                                )
                                existe = cursor.fetchone()

                                if existe:
                                    cursor.execute(
                                        """
                                        UPDATE notas SET Nota = %s
                                        WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'eval' AND Numero = %s
                                        """,
                                        (nota_valor, id_est, id_carga, periodo_activo, i)
                                    )
                                else:
                                    cursor.execute(
                                        """
                                        INSERT INTO notas (ID_Estudiante, ID_Carga, Periodo, Tipo, Numero, Nota)
                                        VALUES (%s, %s, %s, 'eval', %s, %s)
                                        """,
                                        (id_est, id_carga, periodo_activo, i, nota_valor)
                                    )
                            else:
                                flash(f'Nota inválida para {estudiante["Apellidos"]} {estudiante["Nombres"]} en E{i}', 'danger')
                                conn.rollback()
                                return redirect(url_for('planillas_bp.planilla', id_carga=id_carga))
                        except ValueError:
                            flash(f'Valor inválido para {estudiante["Apellidos"]} {estudiante["Nombres"]} en E{i}', 'danger')
                            conn.rollback()
                            return redirect(url_for('planillas_bp.planilla', id_carga=id_carga))
                    else:
                        cursor.execute(
                            """
                            DELETE FROM notas
                            WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'eval' AND Numero = %s
                            """,
                            (id_est, id_carga, periodo_activo, i)
                        )

                for i in range(1, 11):
                    campo_tarea = f'tarea_{id_est}_{i}'
                    valor_tarea = request.form.get(campo_tarea)

                    if valor_tarea is not None and valor_tarea != '':
                        try:
                            nota_valor = float(valor_tarea)
                            if validar_nota(nota_valor, is_dim):
                                cursor.execute(
                                    """
                                    SELECT ID_Nota FROM notas
                                    WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'tarea' AND Numero = %s
                                    """,
                                    (id_est, id_carga, periodo_activo, i)
                                )
                                existe = cursor.fetchone()

                                if existe:
                                    cursor.execute(
                                        """
                                        UPDATE notas SET Nota = %s
                                        WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'tarea' AND Numero = %s
                                        """,
                                        (nota_valor, id_est, id_carga, periodo_activo, i)
                                    )
                                else:
                                    cursor.execute(
                                        """
                                        INSERT INTO notas (ID_Estudiante, ID_Carga, Periodo, Tipo, Numero, Nota)
                                        VALUES (%s, %s, %s, 'tarea', %s, %s)
                                        """,
                                        (id_est, id_carga, periodo_activo, i, nota_valor)
                                    )
                            else:
                                flash(f'Nota inválida para {estudiante["Apellidos"]} {estudiante["Nombres"]} en T{i}', 'danger')
                                conn.rollback()
                                return redirect(url_for('planillas_bp.planilla', id_carga=id_carga))
                        except ValueError:
                            flash(f'Valor inválido para {estudiante["Apellidos"]} {estudiante["Nombres"]} en T{i}', 'danger')
                            conn.rollback()
                            return redirect(url_for('planillas_bp.planilla', id_carga=id_carga))
                    else:
                        cursor.execute(
                            """
                            DELETE FROM notas
                            WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'tarea' AND Numero = %s
                            """,
                            (id_est, id_carga, periodo_activo, i)
                        )

                campo_obs = f'observacion_{id_est}'
                campo_fallas = f'fallas_{id_est}'

                observacion = (request.form.get(campo_obs) or '').strip()
                fallas_raw = (request.form.get(campo_fallas) or '0').strip()

                try:
                    fallas = int(fallas_raw) if fallas_raw else 0
                    if fallas < 0:
                        fallas = 0
                except ValueError:
                    fallas = 0

                guardar_observador_periodo(cursor, id_est, periodo_activo, observacion, fallas)

            indicadores_lista = [request.form.get(f'indicador{i}', '').strip() for i in range(1, 11)]
            guardar_indicadores_planilla(cursor, id_carga, periodo_activo, indicadores_lista)

            conn.commit()
            flash('Notas, indicadores, fallas y observaciones guardados exitosamente', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Error al guardar: {str(e)}', 'danger')

        return redirect(url_for('planillas_bp.planilla', id_carga=id_carga))

    cursor.close()
    conn.close()

    return render_template(
        'planilla.html',
        info=info,
        estudiantes=estudiantes,
        id_carga=id_carga,
        is_dim=is_dim,
        periodo_activo=periodo_activo,
        indicadores_grupo=indicadores_grupo
    )


@planillas_bp.route('/planilla/exportar_excel/<int:id_carga>')
@login_required
@docente_required
def exportar_excel(id_carga):
    conn = get_db_connection()
    cursor = conn.cursor()

    id_docente = get_docente_id_from_user(cursor, session['user_id'])
    if not id_docente:
        conn.close()
        flash('No se encontró información del docente', 'danger')
        return redirect(url_for('auth_bp.panel_docente'))

    cursor.execute(
        """
        SELECT c.ID_Carga, c.ID_Grupo, c.ID_Materia,
               m.Nombre_Materia, g.Nombre_Grupo
        FROM cargas_academicas c
        JOIN materias m ON c.ID_Materia = m.ID_Materia
        JOIN grupos g ON c.ID_Grupo = g.ID_Grupo
        WHERE c.ID_Carga = %s AND c.ID_Docente = %s
        """,
        (id_carga, id_docente)
    )

    info = cursor.fetchone()
    if not info:
        flash('No tiene acceso a esta planilla', 'danger')
        return redirect(url_for('auth_bp.panel_docente'))

    is_dim = es_materia_dim(info['Nombre_Materia'])
    periodo_activo = get_periodo_activo()

    cursor.execute(
        """
        SELECT ID_Estudiante, Nombres, Apellidos
        FROM estudiantes
        WHERE ID_Grupo = %s
        ORDER BY Apellidos, Nombres
        """,
        (info['ID_Grupo'],)
    )

    estudiantes = cursor.fetchall()
    indicadores_grupo = obtener_indicadores_planilla(cursor, id_carga, periodo_activo)

    datos = []
    for estudiante in estudiantes:
        obs = obtener_observador_periodo(cursor, estudiante['ID_Estudiante'], periodo_activo)

        fila = {
            'ID_Estudiante': estudiante['ID_Estudiante'],
            'Apellidos': estudiante['Apellidos'],
            'Nombres': estudiante['Nombres'],
            'Fallas': obs['fallas'],
            'Observacion': obs['observacion']
        }

        cursor.execute(
            """
            SELECT Tipo, Numero, Nota
            FROM notas
            WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s
            """,
            (estudiante['ID_Estudiante'], id_carga, periodo_activo)
        )

        notas = cursor.fetchall()
        notas_dict = {(n['Tipo'], n['Numero']): n['Nota'] for n in notas}

        for i in range(1, 11):
            nota = notas_dict.get(('eval', i))
            fila[f'Eval_{i}'] = nota if nota is not None else ''

        for i in range(1, 11):
            nota = notas_dict.get(('tarea', i))
            fila[f'Tarea_{i}'] = nota if nota is not None else ''

        datos.append(fila)

    conn.close()

    df = pd.DataFrame(datos)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Notas', index=False)
        aplicar_formato_excel(writer, 'Notas')

        df_indicadores = pd.DataFrame([{f'Indicador_{i}': indicadores_grupo[i - 1] for i in range(1, 11)}])
        df_indicadores.to_excel(writer, sheet_name='Indicadores', index=False)
        aplicar_formato_excel(writer, 'Indicadores')

        instrucciones = pd.DataFrame({
            'Instrucción': [
                '📚 INSTRUCCIONES PARA DILIGENCIAR LA PLANILLA',
                '',
                '1. Complete las notas en Eval_1 a Eval_10 y Tarea_1 a Tarea_10',
                '2. Complete Fallas y Observacion para cada estudiante',
                f'3. Tipo de materia: {"Preescolar (DIM) - Solo notas 3, 4 o 5" if is_dim else "Educación Básica - Notas 0 a 5"}',
                '4. Los indicadores están en la hoja Indicadores y son para todo el grupo',
                '5. No modifique los ID_Estudiante ni los nombres',
                '',
                f'📖 Materia: {info["Nombre_Materia"]}',
                f'👥 Grupo: {info["Nombre_Grupo"]}',
                f'📅 Período: {periodo_activo} de 4',
            ]
        })
        instrucciones.to_excel(writer, sheet_name='Instrucciones', index=False)
        aplicar_formato_excel(writer, 'Instrucciones')

    output.seek(0)

    nombre_archivo = f"planilla_{info['Nombre_Materia']}_{info['Nombre_Grupo']}_P{periodo_activo}.xlsx"
    nombre_archivo = nombre_archivo.replace(' ', '_').replace('/', '_')

    return send_file(
        output,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@planillas_bp.route('/planilla/exportar_todas_materias')
@login_required
@docente_required
def exportar_todas_materias():
    conn = get_db_connection()
    cursor = conn.cursor()

    id_docente_db = get_docente_id_from_user(cursor, session['user_id'])

    if not id_docente_db:
        conn.close()
        flash('No se encontró información del docente', 'danger')
        return redirect(url_for('auth_bp.panel_docente'))

    cursor.execute(
        """
        SELECT c.ID_Carga, c.ID_Grupo, c.ID_Materia,
               m.Nombre_Materia, g.Nombre_Grupo, m.Es_DIM
        FROM cargas_academicas c
        JOIN materias m ON c.ID_Materia = m.ID_Materia
        JOIN grupos g ON c.ID_Grupo = g.ID_Grupo
        WHERE c.ID_Docente = %s
        ORDER BY g.Nombre_Grupo, m.Nombre_Materia
        """,
        (id_docente_db,)
    )

    cargas = cursor.fetchall()

    if not cargas:
        flash('No tiene materias asignadas', 'warning')
        return redirect(url_for('auth_bp.panel_docente'))

    periodo_activo = get_periodo_activo()

    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    resumen = []
    indicadores_resumen = []

    for carga in cargas:
        resumen.append({
            'Materia': carga['Nombre_Materia'],
            'Grupo': carga['Nombre_Grupo'],
            'Tipo': 'DIM (Preescolar)' if carga['Es_DIM'] else 'Educación Básica'
        })
        indicadores = obtener_indicadores_planilla(cursor, carga['ID_Carga'], periodo_activo)
        fila_ind = {'Materia': carga['Nombre_Materia'], 'Grupo': carga['Nombre_Grupo']}
        for i in range(1, 11):
            fila_ind[f'Indicador_{i}'] = indicadores[i - 1]
        indicadores_resumen.append(fila_ind)

    df_resumen = pd.DataFrame(resumen)
    df_resumen.to_excel(writer, sheet_name='00_RESUMEN_GENERAL', index=False)
    aplicar_formato_excel(writer, '00_RESUMEN_GENERAL')

    if indicadores_resumen:
        pd.DataFrame(indicadores_resumen).to_excel(writer, sheet_name='INDICADORES', index=False)
        aplicar_formato_excel(writer, 'INDICADORES')

    instrucciones = pd.DataFrame({
        'Instrucción': [
            '📚 INSTRUCCIONES GENERALES',
            '',
            'Este archivo contiene TODAS las materias que usted dicta.',
            'Cada hoja contiene notas, fallas y observación del período.',
            'Los indicadores están en la hoja INDICADORES y aplican por materia y grupo.',
            '',
            f'📅 Período actual: {periodo_activo} de 4'
        ]
    })
    instrucciones.to_excel(writer, sheet_name='INSTRUCCIONES', index=False)
    aplicar_formato_excel(writer, 'INSTRUCCIONES')

    for carga in cargas:
        cursor.execute(
            """
            SELECT ID_Estudiante, Nombres, Apellidos
            FROM estudiantes
            WHERE ID_Grupo = %s
            ORDER BY Apellidos, Nombres
            """,
            (carga['ID_Grupo'],)
        )

        estudiantes = cursor.fetchall()

        datos = []
        for estudiante in estudiantes:
            obs = obtener_observador_periodo(cursor, estudiante['ID_Estudiante'], periodo_activo)

            fila = {
                'ID_Estudiante': estudiante['ID_Estudiante'],
                'Apellidos': estudiante['Apellidos'],
                'Nombres': estudiante['Nombres'],
                'Fallas': obs['fallas'],
                'Observacion': obs['observacion']
            }

            cursor.execute(
                """
                SELECT Tipo, Numero, Nota
                FROM notas
                WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s
                """,
                (estudiante['ID_Estudiante'], carga['ID_Carga'], periodo_activo)
            )

            notas = cursor.fetchall()
            notas_dict = {(n['Tipo'], n['Numero']): n['Nota'] for n in notas}

            for i in range(1, 11):
                nota = notas_dict.get(('eval', i))
                fila[f'Eval_{i}'] = nota if nota is not None else ''

            for i in range(1, 11):
                nota = notas_dict.get(('tarea', i))
                fila[f'Tarea_{i}'] = nota if nota is not None else ''

            datos.append(fila)

        sheet_name = f"{carga['Nombre_Materia']}_{carga['Nombre_Grupo']}"[:31]
        sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('%s', '_').replace('*', '_').replace('[', '_').replace(']', '_')

        df = pd.DataFrame(datos)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        aplicar_formato_excel(writer, sheet_name)

    conn.close()
    writer.close()
    output.seek(0)

    from datetime import datetime
    fecha = datetime.now().strftime('%Y%m%d')
    nombre_archivo = f"planilla_completa_docente_{fecha}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@planillas_bp.route('/planilla/importar_excel/<int:id_carga>', methods=['POST'])
@login_required
@docente_required
def importar_excel(id_carga):
    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo', 'danger')
        return redirect(url_for('planillas_bp.planilla', id_carga=id_carga))

    archivo = request.files['archivo_excel']

    if archivo.filename == '':
        flash('No se seleccionó ningún archivo', 'danger')
        return redirect(url_for('planillas_bp.planilla', id_carga=id_carga))

    if not archivo.filename.endswith(('.xlsx', '.xls')):
        flash('Formato no soportado. Use archivos Excel (.xlsx o .xls)', 'danger')
        return redirect(url_for('planillas_bp.planilla', id_carga=id_carga))

    try:
        excel_file = pd.ExcelFile(archivo)

        if 'Notas' in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name='Notas')
        else:
            df = pd.read_excel(excel_file, sheet_name=0)

        indicadores_excel = ['' for _ in range(10)]
        if 'Indicadores' in excel_file.sheet_names:
            df_ind = pd.read_excel(excel_file, sheet_name='Indicadores')
            if not df_ind.empty:
                fila_ind = df_ind.iloc[0]
                for i in range(1, 11):
                    col = f'Indicador_{i}'
                    if col in df_ind.columns and pd.notna(fila_ind[col]):
                        indicadores_excel[i - 1] = str(fila_ind[col]).strip()

        columnas_requeridas = ['ID_Estudiante', 'Apellidos', 'Nombres']
        for col in columnas_requeridas:
            if col not in df.columns:
                flash(f'El archivo no tiene la columna requerida: {col}', 'danger')
                return redirect(url_for('planillas_bp.planilla', id_carga=id_carga))

        conn = get_db_connection()
        cursor = conn.cursor()

        id_docente = get_docente_id_from_user(cursor, session['user_id'])
        if not id_docente:
            conn.close()
            flash('No se encontró información del docente', 'danger')
            return redirect(url_for('auth_bp.panel_docente'))

        cursor.execute(
            """
            SELECT c.ID_Carga, m.Nombre_Materia
            FROM cargas_academicas c
            JOIN materias m ON c.ID_Materia = m.ID_Materia
            WHERE c.ID_Carga = %s AND c.ID_Docente = %s
            """,
            (id_carga, id_docente)
        )

        info = cursor.fetchone()
        if not info:
            flash('No tiene acceso a esta planilla', 'danger')
            return redirect(url_for('auth_bp.panel_docente'))

        is_dim = es_materia_dim(info['Nombre_Materia'])
        periodo_activo = get_periodo_activo()

        exitosos = 0
        errores = []

        for idx, row in df.iterrows():
            try:
                id_estudiante = int(row['ID_Estudiante'])

                cursor.execute(
                    """
                    SELECT ID_Estudiante FROM estudiantes
                    WHERE ID_Estudiante = %s AND ID_Grupo = (
                        SELECT ID_Grupo FROM cargas_academicas WHERE ID_Carga = %s
                    )
                    """,
                    (id_estudiante, id_carga)
                )

                if not cursor.fetchone():
                    errores.append(f"Fila {idx + 2}: Estudiante ID {id_estudiante} no válido")
                    continue

                for i in range(1, 11):
                    col_eval = f'Eval_{i}'
                    if col_eval in df.columns:
                        valor = row[col_eval]
                        if pd.notna(valor) and valor != '':
                            try:
                                nota = float(valor)
                                if validar_nota(nota, is_dim):
                                    cursor.execute(
                                        """
                                        SELECT ID_Nota FROM notas
                                        WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'eval' AND Numero = %s
                                        """,
                                        (id_estudiante, id_carga, periodo_activo, i)
                                    )
                                    existe = cursor.fetchone()

                                    if existe:
                                        cursor.execute(
                                            """
                                            UPDATE notas SET Nota = %s
                                            WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'eval' AND Numero = %s
                                            """,
                                            (nota, id_estudiante, id_carga, periodo_activo, i)
                                        )
                                    else:
                                        cursor.execute(
                                            """
                                            INSERT INTO notas (ID_Estudiante, ID_Carga, Periodo, Tipo, Numero, Nota)
                                            VALUES (%s, %s, %s, 'eval', %s, %s)
                                            """,
                                            (id_estudiante, id_carga, periodo_activo, i, nota)
                                        )
                                else:
                                    errores.append(f"Fila {idx + 2}: Nota inválida en Eval_{i}")
                            except ValueError:
                                errores.append(f"Fila {idx + 2}: Valor inválido en Eval_{i}")
                        else:
                            cursor.execute(
                                """
                                DELETE FROM notas
                                WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'eval' AND Numero = %s
                                """,
                                (id_estudiante, id_carga, periodo_activo, i)
                            )

                for i in range(1, 11):
                    col_tarea = f'Tarea_{i}'
                    if col_tarea in df.columns:
                        valor = row[col_tarea]
                        if pd.notna(valor) and valor != '':
                            try:
                                nota = float(valor)
                                if validar_nota(nota, is_dim):
                                    cursor.execute(
                                        """
                                        SELECT ID_Nota FROM notas
                                        WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'tarea' AND Numero = %s
                                        """,
                                        (id_estudiante, id_carga, periodo_activo, i)
                                    )
                                    existe = cursor.fetchone()

                                    if existe:
                                        cursor.execute(
                                            """
                                            UPDATE notas SET Nota = %s
                                            WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'tarea' AND Numero = %s
                                            """,
                                            (nota, id_estudiante, id_carga, periodo_activo, i)
                                        )
                                    else:
                                        cursor.execute(
                                            """
                                            INSERT INTO notas (ID_Estudiante, ID_Carga, Periodo, Tipo, Numero, Nota)
                                            VALUES (%s, %s, %s, 'tarea', %s, %s)
                                            """,
                                            (id_estudiante, id_carga, periodo_activo, i, nota)
                                        )
                                else:
                                    errores.append(f"Fila {idx + 2}: Nota inválida en Tarea_{i}")
                            except ValueError:
                                errores.append(f"Fila {idx + 2}: Valor inválido en Tarea_{i}")
                        else:
                            cursor.execute(
                                """
                                DELETE FROM notas
                                WHERE ID_Estudiante = %s AND ID_Carga = %s AND Periodo = %s AND Tipo = 'tarea' AND Numero = %s
                                """,
                                (id_estudiante, id_carga, periodo_activo, i)
                            )

                observacion = str(row['Observacion']).strip() if 'Observacion' in df.columns and pd.notna(row['Observacion']) else ''
                fallas = 0
                if 'Fallas' in df.columns and pd.notna(row['Fallas']):
                    try:
                        fallas = int(row['Fallas'])
                        if fallas < 0:
                            fallas = 0
                    except ValueError:
                        fallas = 0

                guardar_observador_periodo(cursor, id_estudiante, periodo_activo, observacion, fallas)

                exitosos += 1

            except Exception as e:
                errores.append(f"Fila {idx + 2}: {str(e)}")

        guardar_indicadores_planilla(cursor, id_carga, periodo_activo, indicadores_excel)
        conn.commit()
        conn.close()

        if exitosos > 0:
            flash(f'✅ {exitosos} estudiantes procesados exitosamente', 'success')

        if errores:
            flash(f'⚠️ {len(errores)} errores encontrados. Primeros 5: {", ".join(errores[:5])}', 'warning')

    except Exception as e:
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')

    return redirect(url_for('planillas_bp.planilla', id_carga=id_carga))
